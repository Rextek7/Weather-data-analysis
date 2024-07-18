import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from analyse import df_moscow, df_sankt_peterburg, df_yalta
import numpy as np


def calculate_stats(df):
    """
    Рассчитывает статистические метрики для указанного DataFrame.

    Args:
        df: DataFrame с данными о погоде.

    Returns:
        stats: Словарь, содержащий статистические метрики.
    """
    stats = {
        'Средняя дневная температура': np.mean(df['day_temp']),
        'Средняя ночная температура': np.mean(df['night_temp']),
        'Среднее количество осадков': np.mean(df['precipitation']),
        'Средняя скорость ветра': np.mean(df['wind_m_s']),
        'Среднее максимальное давление': np.mean(df['max_pressure']),
        'Среднее минимальное давление': np.mean(df['min_pressure']),
    }
    return stats


def create_empty_excel(filename):
    """Создает новый файл Excel с пустым листом."""
    with pd.ExcelWriter(filename) as writer:
        pd.DataFrame().to_excel(writer, sheet_name='Sheet1')
    return filename


def save_data_to_excel(writer, sheet_name, data, columns):
    """
    Сохраняет данные в указанный лист Excel файла и настраивает ширину колонок.

    Args:
        writer: Объект ExcelWriter.
        sheet_name: Название листа Excel.
        data: Данные для сохранения.
        columns: Названия столбцов.
    """
    df = pd.DataFrame(data, columns=columns)
    startrow = writer.sheets[sheet_name].max_row - 1  # Найти первую пустую строку
    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=startrow)

    # Настройка ширины колонок и выравнивания текста по центру
    worksheet = writer.sheets[sheet_name]
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            cell.alignment = Alignment(horizontal='left', vertical='center')
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    for row in worksheet.rows:
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')


def save_chart_to_excel(writer, sheet_name, fig, image_path):
    """
    Сохраняет диаграмму в виде изображения и вставляет его на указанный лист Excel файла.

    Args:
        writer: Объект ExcelWriter.
        sheet_name: Название листа Excel.
        fig: Объект matplotlib.figure.Figure с построенной диаграммой.
        image_path: Путь для сохранения изображения.
    """
    fig.savefig(image_path, format='png')
    plt.close(fig)  # Закрываем фигуру, чтобы избежать отображения пустой диаграммы
    workbook = writer.book
    worksheet = workbook[sheet_name]
    img = Image(image_path)
    worksheet.add_image(img, 'A5')


def main():
    # Создаем новый файл Excel
    filename = 'results_part2.xlsx'
    create_empty_excel(filename)

    # Рассчитываем статистические метрики для каждого города
    stats_moscow = calculate_stats(df_moscow)
    stats_sankt_peterburg = calculate_stats(df_sankt_peterburg)
    stats_yalta = calculate_stats(df_yalta)

    with (pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer):

        # Добавляем данные о погоде для каждого города в отдельные листы
        sheet_name = 'Данные в городах'
        workbook = writer.book
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        columns = ['День', 'Дневная температура (°C)', 'Ночная температура (°C)', 'Осадки (мм)', 'Скорость ветра (м/с)',
                   'Максимальное давление (мм рт. ст.)', 'Минимальное давление (мм рт. ст.)', 'Дата             ']

        city_ms = ['Москва']
        data_ms = [city_ms] + [columns] + df_moscow.reset_index().values.tolist()

        city_sp = ['Санкт-Петербург']
        data_sp = [''] + [city_sp] + [columns] + df_sankt_peterburg.reset_index().values.tolist()

        city_yl = ['Ялта']
        data_yl = [''] + [city_yl] + [columns] + df_yalta.reset_index().values.tolist()

        res = data_ms + data_sp + data_yl
        save_data_to_excel(writer, sheet_name, res, columns)

        # Создаем лист для анализа погоды
        sheet_name = 'Анализ погоды'
        workbook = writer.book
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        # Добавляем статистику для каждого города
        stats_data = [
            ['Москва', stats_moscow['Средняя дневная температура'], stats_moscow['Средняя ночная температура'],
             stats_moscow['Среднее количество осадков'], stats_moscow['Средняя скорость ветра'],
             stats_moscow['Среднее максимальное давление'], stats_moscow['Среднее минимальное давление']],
            ['Санкт-Петербург', stats_sankt_peterburg['Средняя дневная температура'],
             stats_sankt_peterburg['Средняя ночная температура'], stats_sankt_peterburg['Среднее количество осадков'],
             stats_sankt_peterburg['Средняя скорость ветра'], stats_sankt_peterburg['Среднее максимальное давление'],
             stats_sankt_peterburg['Среднее минимальное давление']],
            ['Ялта', stats_yalta['Средняя дневная температура'], stats_yalta['Средняя ночная температура'],
             stats_yalta['Среднее количество осадков'], stats_yalta['Средняя скорость ветра'],
             stats_yalta['Среднее максимальное давление'], stats_yalta['Среднее минимальное давление']]
        ]
        columns = ['Город', 'Средняя дневная температура', 'Средняя ночная температура', 'Среднее количество осадков',
                   'Средняя скорость ветра', 'Среднее максимальное давление', 'Среднее минимальное давление']

        data_col = [columns] + stats_data
        save_data_to_excel(writer, sheet_name, data_col, columns)

        # Построим графики
        fig, axs = plt.subplots(3, 2, figsize=(14, 10))

        # График дневной температуры
        axs[0, 0].plot(df_moscow['date'], df_moscow['day_temp'], label='Москва')
        axs[0, 0].plot(df_sankt_peterburg['date'], df_sankt_peterburg['day_temp'], label='Санкт-Петербург')
        axs[0, 0].plot(df_yalta['date'], df_yalta['day_temp'], label='Ялта')
        axs[0, 0].set_xlabel('Дата')
        axs[0, 0].set_ylabel('Дневная температура (°C)')
        axs[0, 0].set_title('Дневная температура')
        axs[0, 0].legend()

        # График ночной температуры
        axs[0, 1].plot(df_moscow['date'], df_moscow['night_temp'], label='Москва')
        axs[0, 1].plot(df_sankt_peterburg['date'], df_sankt_peterburg['night_temp'], label='Санкт-Петербург')
        axs[0, 1].plot(df_yalta['date'], df_yalta['night_temp'], label='Ялта')
        axs[0, 1].set_xlabel('Дата')
        axs[0, 1].set_ylabel('Ночная температура (°C)')
        axs[0, 1].set_title('Ночная температура')
        axs[0, 1].legend()

        # График осадков
        axs[1, 0].plot(df_moscow['date'], df_moscow['precipitation'], label='Москва')
        axs[1, 0].plot(df_sankt_peterburg['date'], df_sankt_peterburg['precipitation'], label='Санкт-Петербург')
        axs[1, 0].plot(df_yalta['date'], df_yalta['precipitation'], label='Ялта')
        axs[1, 0].set_xlabel('Дата')
        axs[1, 0].set_ylabel('Осадки (мм)')
        axs[1, 0].set_title('Осадки')
        axs[1, 0].legend()

        # График скорости ветра
        axs[1, 1].plot(df_moscow['date'], df_moscow['wind_m_s'], label='Москва')
        axs[1, 1].plot(df_sankt_peterburg['date'], df_sankt_peterburg['wind_m_s'], label='Санкт-Петербург')
        axs[1, 1].plot(df_yalta['date'], df_yalta['wind_m_s'], label='Ялта')
        axs[1, 1].set_xlabel('Дата')
        axs[1, 1].set_ylabel('Скорость ветра (м/с)')
        axs[1, 1].set_title('Скорость ветра')
        axs[1, 1].legend()

        # График максимального давления
        axs[2, 0].plot(df_moscow['date'], df_moscow['max_pressure'], label='Москва')
        axs[2, 0].plot(df_sankt_peterburg['date'], df_sankt_peterburg['max_pressure'], label='Санкт-Петербург')
        axs[2, 0].plot(df_yalta['date'], df_yalta['max_pressure'], label='Ялта')
        axs[2, 0].set_xlabel('Дата')
        axs[2, 0].set_ylabel('Максимальное давление (мм рт. ст.)')
        axs[2, 0].set_title('Максимальное давление')
        axs[2, 0].legend()

        # График минимального давления
        axs[2, 1].plot(df_moscow['date'], df_moscow['min_pressure'], label='Москва')
        axs[2, 1].plot(df_sankt_peterburg['date'], df_sankt_peterburg['min_pressure'], label='Санкт-Петербург')
        axs[2, 1].plot(df_yalta['date'], df_yalta['min_pressure'], label='Ялта')
        axs[2, 1].set_xlabel('Дата')
        axs[2, 1].set_ylabel('Минимальное давление (мм рт. ст.)')
        axs[2, 1].set_title('Минимальное давление')
        axs[2, 1].legend()

        plt.tight_layout()
        image_path = 'weather_analysis_chart.png'
        save_chart_to_excel(writer, sheet_name, fig, image_path)

        # Присваиваем название и создаем новую вкладку
        sheet_name = 'Вывод'
        workbook = writer.book
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)

        columns = ['Параметр', 'Вывод']

        data = [
            columns,
            ['Дневная и ночная температура:', 'Москва и Санкт-Петербург имеют более выраженные колебания температуры в '
                                              'течение дня и ночи по сравнению с Ялтой.\n'
                                              ' В Ялте ночные температуры более стабильны и выше, чем в '
                                              'Москве и Санкт-Петербурге.'],
            ['Осадки:', 'В Москве и Санкт-Петербурге наблюдаются периоды значительных осадков, в то время как в '
                        'Ялте осадки практически отсутствуют в течение рассматриваемого периода, за '
                        'исключением нескольких дней.'],
            ['Скорость ветра:', 'Средняя скорость ветра выше в Ялте, особенно в начале периода.\n В Санкт-Петербурге '
                                'также отмечаются высокие значения скорости ветра в некоторые дни.'],
            ['Давление:', 'Давление в Москве и Санкт-Петербурге варьируется более значительно, '
                          'чем в Ялте, где оно более стабильное.']

        ]
        # Сохраняем результат с помощью функции
        save_data_to_excel(writer, sheet_name, data, columns)
    # Удаляем первую вкладку
    workbook = load_workbook(filename)
    workbook.remove(workbook['Sheet1'])
    # Сохраняем изменения и закрываем файл
    workbook.save(filename)
    workbook.close()
    print("Файл Excel готов")

if __name__ == "__main__":
    main()
